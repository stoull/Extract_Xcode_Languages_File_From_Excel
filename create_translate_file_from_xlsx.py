from openpyxl import load_workbook
import re
from enum import Enum
from pathlib import Path
import glob

class Language(Enum):
	chinese_simplifed = 5
	english_en = 6
	italian_it = 7
	polish_pl = 8
	putch_nl = 9
	german_de = 10
	hungarian_hu = 11
	portuguese_pt = 12
	chinese_hongkong = 13
	chinese_traditional = 14
	spanish_es = 15
	korean_ko = 16
	french_fr = 17

	# 获取对应语言的String识别符
	def getIdentifier(language):
		fileName="chinese_simplifed"
		localizedName="中文简体"
		if language == Language.chinese_simplifed:
			fileName="chinese_simplifed"
			localizedName="中文简体"
		elif language == Language.english_en:
			fileName="english_en"
			localizedName="英文"
		elif language == Language.italian_it:
			fileName="italian_it"
			localizedName="意大利文"
		elif language == Language.polish_pl:
			fileName="polish_pl"
			localizedName="波兰语"
		elif language == Language.putch_nl:
			fileName="putch_nl"
			localizedName="荷兰语"
		elif language == Language.german_de:
			fileName="german_de"
			localizedName="德语"
		elif language == Language.hungarian_hu:
			fileName="hungarian_hu"
			localizedName="匈牙利语"
		elif language == Language.portuguese_pt:
			fileName="portuguese_pt"
			localizedName="葡萄牙语"
		elif language == Language.chinese_hongkong:
			fileName="chinese_hongkong"
			localizedName="中文香港繁体"
		elif language == Language.chinese_traditional:
			fileName="chinese_traditional"
			localizedName="中文繁体"
		elif language == Language.spanish_es:
			fileName="spanish_es"
			localizedName="西班牙语"
		elif language == Language.korean_ko:
			fileName="korean_ko"
			localizedName="韩语"
		elif language == Language.french_fr:
			fileName="french_fr"
			localizedName="法语"
		else:
			fileName="chinese_simplifed"
			localizedName="中文简体"
		return fileName,localizedName

# 从base文件中取值。返回的为dictionary
def getBaseInfoFromWatchLanguageStringFile():
	baseKeysValues = {}
	with open('Base.strings', 'r') as file:
		for line in file:
			result = re.match(r'"(.*)" *?= *?"(.*)";', line)
			if result:
				key = result.group(1)
				value = result.group(2)
				if len(key)>0:
					baseKeysValues[key]=value
	return baseKeysValues

# 检查生成的结果是否有缺失。即Base.string中所有的key在结果的.string文件中都存在
def result_check():
	print("======== 文件完整性检查结果	======== ")
	resultStringPaths=glob.glob("./Result/*.strings")
	allLogMessages=[]
	if len(resultStringPaths)>0:
		#查找没有匹配上的行只要一个loop就可以
		isNoMatchLinsExcuted = False
		for resultFilePath in resultStringPaths:
			reStrFile = re.match(r'^\./.*/(.*).strings$', resultFilePath)
			if reStrFile:
				currentLanguage = reStrFile.group(1)
			with open("Base.strings", 'r') as referFile:
				with open(resultFilePath, 'r') as resultFile:
					isAllExit = True
					for line1 in referFile:
						re1 = re.match(r'(".*") *?= *?(".*");', line1)
						isExit = False
						resultFile.seek(0)
						if re1:
							key1 = re1.group(1)
							value1 = re1.group(2)
							for line2 in resultFile:
								re2 = re.match(r'(".*") *?= *?(".*");', line2)
								if re2:
									key2 = re2.group(1)
									value2 = re2.group(2)
									if key1 == key2:
										isExit = True
										break
						else:
							isExit = True
							if isNoMatchLinsExcuted == False:
								message=f"watos_base文件中没有作处理的行 {line1}"
								print(message)
								allLogMessages.append(message)
						if isExit == False:
							if len(line1) > 0:
								isAllExit = False
								errorMessage=f"{line1} 在文件:{currentLanguage} 中缺失！"
								print(errorMessage)
								allLogMessages.append(errorMessage)
					if isAllExit == False:
						logMessage = f'{currentLanguage}.strings has errors! Check it!'
						allLogMessages.append(logMessage)
						allLogMessages.append('\n' + '==================' + '\n')
						print(logMessage)
					else:
						message=f"Check {resultFilePath}: Success! 无缺失key-value行，但不保证所有都正确翻译!，请查看../log.txt文件，查看哪些没有翻译。"
						print(message)
						allLogMessages.append(message)
						allLogMessages.append('\n' + '==================' + '\n')
			isNoMatchLinsExcuted = True
	else:
		logMessage="未找到任何翻译文件！"
		print(logMessage)
		allLogMessages.append(logMessage)
		allLogMessages.append('\n' + '==================' + '\n')
	with open('./result_check_log.txt', mode='wt', encoding='utf-8') as logFile:
		logFile.write('\n'.join(allLogMessages))
	print("======== 文件完整性结束	======== ")

def getValueExistIndex(valueString, targetList):
		try:
			exitIndex = targetList.index(valueString)
			return exitIndex
		except Exception as e:
			return False

def getStringsFromXlsxFile():
	wb = load_workbook('source_translate.xlsx')
	transSheet = ""
	for sheet in wb:
		stitle = sheet.title
		if stitle == "翻译（新修改和新增用红色标记）":
			print(f"selected sheet: {stitle}")
			transSheet = sheet
			break

	# 获取 安卓key值

	# 获取 苹果key值
	appleKey_Keys=[]
	loopIndex = 0
	for row in transSheet.iter_rows(min_row=1, max_row=3663, min_col=4, max_col=4, values_only=True):
		rowValue = row[0]
		loopIndex+=1
		if rowValue:
			rowValue = str(rowValue)
			# 移除 xlsx 文件单元格中开始和结束误输入的空格
			rowValue = re.sub('^[ ]*|[ ]*$', '', rowValue)
			appleKey_Keys.append(rowValue)
		else:
			appleKey_Keys.append("")
			# print(f"苹果key值 列在第： {loopIndex}行为空！")

	# 获取 中文（校对）行的值
	baseChinese_Simplifed=[]
	loopIndex = 0
	for row in transSheet.iter_rows(min_row=1, max_row=3663, min_col=5, max_col=5, values_only=True):
		rowValue = row[0]
		loopIndex+=1
		if rowValue:
			rowValue = str(rowValue)
			# 移除 xlsx 文件单元格中开始和结束误输入的空格及其它的特殊字符
			rowValue = re.sub('^[ ]*|[ ]*$', '', rowValue)
			baseChinese_Simplifed.append(rowValue)
		else:
			baseChinese_Simplifed.append("")
			# print(f"中文（校对）列在第： {loopIndex}行为空！")

	# with open("./appleKey_Keys.txt", mode='wt', encoding='utf-8') as the_file:
	# 	 the_file.write('\n'.join(appleKey_Keys))

	# with open("./baseChinese_Simplifed.txt", mode='wt', encoding='utf-8') as the_file:
	# 	 the_file.write('\n'.join(baseChinese_Simplifed))

	# 获取基准文件的key和简体中文Value
	baseKeyValues = getBaseInfoFromWatchLanguageStringFile()

	# 收集收有语言信息，对应的顺序为枚举Language中的顺序
	allResutls=[[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[]]
	allUnTranslated=[[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[]]

	# 收集所有对应语言下未翻译的内容
	allNotFoundResult= [[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[],[]]
	for index in range(0,17,1):
		try:
			lan = Language(index)
			allNotFoundResult[index].append(f"//====== 未翻译的{lan.getIdentifier()[1]}: ======")
		except Exception as e:
			allNotFoundResult[index].append(f"//====== 未知语言类型{index} ======")
			continue

	# 所包含的语言
	requireLans = [Language.chinese_simplifed, Language.english_en, Language.italian_it,
				Language.polish_pl, Language.putch_nl, Language.german_de,
				Language.hungarian_hu, Language.portuguese_pt, Language.chinese_hongkong,
				Language.chinese_traditional, Language.spanish_es, Language.korean_ko, Language.french_fr]
	
	# 循环所有的key，进行匹配对应的翻译
	baseKeyItems = baseKeyValues.items()
	if len(baseKeyItems) < 1:
		print("在文件：Base.strings 中，无任何目标翻译行")
	else:
		for key,value in baseKeyItems:
			# 用baseKeyValues中的中文内容，先匹配xlsx文件中苹果key值行，如果没有再匹配 中文（校对）行，并对应该行的其它语言
			exitIndexValueList = appleKey_Keys 
			exitIndex = getValueExistIndex(valueString=key, targetList=exitIndexValueList)
			if exitIndex == False:
				exitIndexValueList = baseChinese_Simplifed
				exitIndex = getValueExistIndex(valueString=value, targetList=exitIndexValueList)
				if exitIndex == False:
					for lan in requireLans:
						resultLine = f"\"{key}\" = \"{value}\";"
						allUnTranslated[lan.value].append(resultLine)
						allNotFoundResult[lan.value].append(f"{resultLine} 翻译文件中未查找到翻译信息！")
					continue

			rowCells = transSheet[exitIndex+1]
			for lan in requireLans:
				cellValue = rowCells[lan.value-1].value
				if cellValue:
					# 移除 xlsx 文件单元格中的换行符等
					cellValue = cellValue.replace('\n','')
					isExit = True
					resultLine = f"\"{key}\" = \"{cellValue}\";"
					allResutls[lan.value].append(resultLine)
				else:
					resultLine = f"\"{key}\" = \"{value}\";"
					allNotFoundResult[lan.value].append(f"{resultLine}")
					allUnTranslated[lan.value].append(resultLine)
					print(f"{resultLine} 在行：{exitIndex}  列：{lan.value} 中没有翻译！")

		totalNotFoundMessages=[]

		for notFound in allNotFoundResult:
			if len(notFound)>1:
				totalNotFoundMessages.extend(notFound)

		exceptionFilePath = f"./log.txt"
		with open(exceptionFilePath, mode='wt', encoding='utf-8') as the_file:
			the_file.write('\n'.join(totalNotFoundMessages))

		for index in range(1,len(allResutls),1):
			resultList = allResutls[index]
			unTransledList = allNotFoundResult[index]
			resultList.extend(unTransledList)
			lanIdentifier = requireLans[index-5].getIdentifier()[0]
			filePath = f"./Result/{lanIdentifier}.strings"
			if len(resultList) > 0:
				with open(filePath, mode='wt', encoding='utf-8') as the_file:
					 the_file.write('\n'.join(resultList))

if __name__ == '__main__':
	getStringsFromXlsxFile()
	result_check()